import gymnasium as gym
import minigrid
import babyai
import os
import json
import numpy as np
import torch
import torch.nn as nn
from stable_baselines3 import PPO
from stable_baselines3.common.monitor import Monitor
from stable_baselines3.common.vec_env import DummyVecEnv, VecNormalize
from stable_baselines3.common.torch_layers import BaseFeaturesExtractor
from stable_baselines3.common.callbacks import CheckpointCallback, BaseCallback
from Reward_Algorithm import *
from Grid_Maze_360_vision import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ImprovedCNN(BaseFeaturesExtractor):
    """
    Same as the one in Training.
    """
    def __init__(self, observation_space, features_dim=256):
        super().__init__(observation_space, features_dim)

        obs_shape = observation_space.shape
        in_channels = obs_shape[0]
        H, W = obs_shape[1], obs_shape[2]

        self.obj_stream = nn.Sequential(
            nn.Conv2d(in_channels, 32, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Conv2d(32, 64, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Conv2d(64, 64, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Flatten(),
        )

        self.coord_stream = nn.Sequential(
            nn.Conv2d(2, 16, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Conv2d(16, 32, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Conv2d(32, 32, kernel_size=3, stride=1, padding=1),
            nn.LeakyReLU(0.1),
            nn.Flatten(),
        )

        cx, cy = W // 2, H // 2
        x_coords = (torch.arange(W).float() - cx) / max(cx, 1)
        y_coords = (torch.arange(H).float() - cy) / max(cy, 1)
        x_grid = x_coords.unsqueeze(0).expand(H, -1)
        y_grid = y_coords.unsqueeze(1).expand(-1, W)
        coord_grid = torch.stack([x_grid, y_grid], dim=0).unsqueeze(0)
        self.register_buffer('coord_grid', coord_grid)

        with torch.no_grad():
            dummy_obj   = torch.zeros(1, in_channels, H, W)
            dummy_coord = torch.zeros(1, 2, H, W)
            n_obj   = self.obj_stream(dummy_obj).shape[1]
            n_coord = self.coord_stream(dummy_coord).shape[1]

        self.fusion = nn.Sequential(
            nn.Linear(n_obj + n_coord, features_dim),
            nn.LayerNorm(features_dim),
            nn.Tanh()
        )

        self._apply_init()

    def _apply_init(self):
        for m in self.modules():
            if isinstance(m, (nn.Conv2d, nn.Linear)):
                nn.init.orthogonal_(m.weight, gain=1.414)
                if m.bias is not None:
                    nn.init.constant_(m.bias, 0)

    def forward(self, observations):
        obs = observations.float() / 10.0
        B = obs.shape[0]
        coords = self.coord_grid.expand(B, -1, -1, -1)
        obj_feat   = self.obj_stream(obs)
        coord_feat = self.coord_stream(coords)
        combined = torch.cat([obj_feat, coord_feat], dim=1)
        return self.fusion(combined)


if "BabyAI-CleanTask-v0" not in gym.envs.registration.registry:
    gym.register(
        id='BabyAI-CleanTask-v0',
        entry_point=CleanTaskEnv,
    )

class FixedSeedWrapper(gym.Wrapper):
    """
    Forces every env.reset() call to use the same fixed max seed. 
    Also resets step_count to 0 to prevent premature episode truncation.
    """
    def __init__(self, env, seed):
        super().__init__(env)
        self.fixed_seed = seed

    def reset(self, **kwargs):
        kwargs['seed'] = self.fixed_seed
        obs, info = self.env.reset(**kwargs)

        base_env = self.env
        while hasattr(base_env, 'env'):
            base_env = base_env.env
        base_env.step_count = 0

        return obs, info

class SuccessTrackingCallback(BaseCallback):
    """
    Records each successful episode along with its relative timestep and episode length.
    Prints a rolling success summary every log_freq steps.
    """
    def __init__(self, log_freq=2000, verbose=0):
        super().__init__(verbose)
        self.log_freq         = log_freq
        self.success_episodes = []
        self.start_timestep   = None 

    def _on_training_start(self) -> None:
        self.start_timestep = self.num_timesteps  

    def _on_step(self) -> bool:
        relative_t = self.num_timesteps - self.start_timestep  

        for info in self.locals.get('infos', []):
            ep_info = info.get('episode')
            if ep_info is not None and ep_info.get('r', 0) >= 50:
                self.success_episodes.append({
                    'timestep': relative_t, 
                    'steps':    ep_info['l'],
                    'reward':   ep_info['r'],
                })

        if relative_t % self.log_freq == 0 and relative_t > 0:
            recent = [
                e for e in self.success_episodes
                if e['timestep'] > relative_t - self.log_freq
            ]
            total = len(self.success_episodes)
            if recent:
                avg_steps = np.mean([e['steps'] for e in recent])
                print(f"  [Step {relative_t}] "
                      f"Last {self.log_freq} steps: {len(recent)} successes | "
                      f"Avg steps: {avg_steps:.1f} | "
                      f"Total successes: {total}")
            else:
                print(f"  [Step {relative_t}] "
                      f"Last {self.log_freq} steps: 0 successes | "
                      f"Total successes: {total}")
        return True

def make_env_phase2(room_size, num_walls, num_pockets,
                    use_control_reward=False, control_scale=2.0,
                    map_seed=2000):
    def _init():
        env = gym.make(
            "BabyAI-CleanTask-v0",
            room_size=room_size,
            num_walls=num_walls,
            num_pockets=num_pockets,
            use_distance_reward=False,
            has_goal=True,
            render_mode="rgb_array"
        )
        env = ImgObsWrapper(env)

        if use_control_reward:
            env = EmpiricalInstrumentalDivergenceWrapper(
                env, scale=control_scale,
            )

        env = FixedSeedWrapper(env, seed=map_seed)
        env = Monitor(env)
        return env
    return _init

def train_phase2(mode, control_scale=2.0, map_seed=2000, total_steps=100000):
    """
    Learning in Testing: fine-tunes agents on a fixed map with has_goal=True.
    - 'random'  mode: initializes a random PPO model as a sanity check baseline.
    - 'baseline' mode: loads the pre-trained baseline checkpoint and continues training.
    - 'control'  mode: loads the pre-trained control checkpoint and continues training.
    """
    print(f"\n{'='*60}")
    print(f"Learning in Testing: {mode.upper()} | has_goal=True | map_seed={map_seed}")
    if mode != "random":
        print(f"Loading pre-trained model: model_{mode}_multimap_final.zip")
    else:
        print(f"Random model: initializing from scratch (sanity check)")
    print(f"{'='*60}\n")
 
    use_control = (mode == "control")
 
    env = DummyVecEnv([
        make_env_phase2(
            room_size=11,
            num_walls=22,
            num_pockets=10,
            use_control_reward=use_control,
            control_scale=control_scale,
            map_seed=map_seed,
        )
    ])
 
    env = VecNormalize(
        env,
        norm_obs=False,
        norm_reward=True,
        gamma=0.99,
    )
 
    policy_kwargs = dict(
        features_extractor_class=ImprovedCNN,
        features_extractor_kwargs=dict(features_dim=256),
        share_features_extractor=False
    )
 
    if mode == "random":
        model = PPO(
            "CnnPolicy",
            env,
            policy_kwargs=policy_kwargs,
            learning_rate=2e-4,
            n_steps=2048,
            batch_size=256,
            n_epochs=10,
            gamma=0.99,
            gae_lambda=0.95,
            clip_range=0.2,
            ent_coef=0.05,
            vf_coef=1.0,
            max_grad_norm=0.5,
            verbose=1,
            tensorboard_log=f"./logs_phase2_{mode}/",
        )
    else:
         # Load pretrained weights; all hyperparameters are preserved
        model = PPO.load(
            f"model_{mode}_multimap_final",
            env=env,
            tensorboard_log=f"./logs_phase2_{mode}/",
        )
 
    # Print map layout info before training begins
    env.reset()
    base_env = env.envs[0]
    while hasattr(base_env, 'env'):
        base_env = base_env.env
    print(f"Map info | agent_pos={tuple(base_env.agent_pos)} | "
          f"ball_pos={tuple(base_env.red_ball.cur_pos)}\n")
 
    success_callback = SuccessTrackingCallback(log_freq=2000)
 
    callbacks = [
        CheckpointCallback(
            save_freq=50000,
            save_path=f'./checkpoints_{mode}_phase2/',
            name_prefix='model'
        ),
        success_callback,
    ]
 
    print(f"Start learning...\n")
    model.learn(
        total_timesteps=total_steps,
        callback=callbacks,
        reset_num_timesteps=False,
        progress_bar=False,
    )
 
    # Save model
    model.save(f"model_{mode}_phase2_final")
    print(f"\nModel saved: model_{mode}_phase2_final.zip")
 
    # Save success log as JSON for later Excel export
    log_path = f"success_log_{mode}_phase2.json"
    with open(log_path, "w") as f:
        json.dump(success_callback.success_episodes, f, indent=2)
    print(f"Success log saved: {log_path}")
 
    # Print step trend
    successes = success_callback.success_episodes
    if successes:
        steps_list = [e['steps'] for e in successes]
        print(f"\nSuccess summary:")
        print(f"  Total successes: {len(successes)}")
        print(f"  Avg steps:       {np.mean(steps_list):.1f}")
        print(f"  Min steps:       {np.min(steps_list)}")
        print(f"  Max steps:       {np.max(steps_list)}")
 
        seg_size = total_steps // 5
        print(f"\n  Step trend (per {seg_size} steps):")
        for seg in range(5):
            seg_start = seg * seg_size
            seg_end   = (seg + 1) * seg_size
            seg_data  = [e['steps'] for e in successes
                         if seg_start <= e['timestep'] < seg_end]
            if seg_data:
                print(f"    {seg_start:>7}-{seg_end:<7}: "
                      f"{len(seg_data):>4} successes | "
                      f"Avg steps: {np.mean(seg_data):.1f}")
            else:
                print(f"    {seg_start:>7}-{seg_end:<7}: no successes")
    else:
        print("\nNo successful episodes during training.")
 
    print(f"\n{'='*60}")
    return model
 
def export_summary_excel(modes, total_steps, log_freq=2000,
                          output_path="phase2_summary.xlsx"):
    logs = {}
    for mode in modes:
        log_path = f"success_log_{mode}_phase2.json"
        if os.path.exists(log_path):
            with open(log_path) as f:
                logs[mode] = json.load(f)
        else:
            print(f"Warning: {log_path} not found, using empty data")
            logs[mode] = []
 
    windows = list(range(log_freq, total_steps + log_freq, log_freq))
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Summary"
 
    # Styles
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    subheader_font = Font(name="Arial", bold=True, size=10)
    data_font      = Font(name="Arial", size=10)
    total_font     = Font(name="Arial", bold=True, size=10)
 
    header_fill    = PatternFill("solid", start_color="1F4E79")
    subheader_fill = PatternFill("solid", start_color="2E75B6")
    alt_fill       = PatternFill("solid", start_color="EBF3FB")
    total_fill     = PatternFill("solid", start_color="D6E4F0")
 
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
 
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    mode_colors = {
        "random":   "C55A11",
        "baseline": "2E75B6",
        "control":  "375623",
    }
 
    # Title row
    n_cols = 1 + len(modes) * 2
    last_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "Phase 2 Training Summary — Successes & Avg Steps per 2000 Timesteps"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28
 
    # Config info row
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (f"Total steps: {total_steps:,}    |    "
                f"Window: {log_freq:,} steps    |    "
                f"Modes: {', '.join(m.capitalize() for m in modes)}")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = left
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6  # spacer
 
    # Column headers rows 4-5
    col_map = {}
    start_col = 2
 
    # Step Window header (spans rows 4-5)
    ws.cell(row=4, column=1, value="Step Window")
    ws.cell(row=4, column=1).font      = subheader_font
    ws.cell(row=4, column=1).fill      = subheader_fill
    ws.cell(row=4, column=1).alignment = center
    ws.cell(row=4, column=1).border    = border
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws.column_dimensions["A"].width = 20
 
    for i, mode in enumerate(modes):
        sc = start_col + i * 2
        col_map[mode] = (sc, sc + 1)
 
        # Group header
        ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=sc+1)
        cell = ws.cell(row=4, column=sc, value=mode.capitalize())
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color=mode_colors.get(mode, "2E75B6"))
        cell.alignment = center
        cell.border    = border
 
        # Sub-headers
        for col, label in [(sc, "Successes"), (sc+1, "Avg Steps")]:
            c = ws.cell(row=5, column=col, value=label)
            c.font      = subheader_font
            c.fill      = PatternFill("solid", start_color="D9E2F3")
            c.alignment = center
            c.border    = border
            ws.column_dimensions[get_column_letter(col)].width = 14
 
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 18
 
    # Data rows
    data_start_row = 6
    for idx, window_end in enumerate(windows):
        window_start = window_end - log_freq
        row  = data_start_row + idx
        fill = alt_fill if idx % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
 
        c = ws.cell(row=row, column=1,
                    value=f"{window_start:,} – {window_end:,}")
        c.font      = data_font
        c.fill      = fill
        c.alignment = center
        c.border    = border
 
        for mode in modes:
            sc, ac = col_map[mode]
            recent    = [e for e in logs[mode]
                         if window_start < e['timestep'] <= window_end]
            count     = len(recent)
            avg_steps = round(np.mean([e['steps'] for e in recent]), 1) if recent else 0.0
 
            for col, val in [(sc, count), (ac, avg_steps)]:
                c = ws.cell(row=row, column=col, value=val)
                c.font      = data_font
                c.fill      = fill
                c.alignment = center
                c.border    = border
 
        ws.row_dimensions[row].height = 16
 
    # Total row
    total_row = data_start_row + len(windows)
    ws.row_dimensions[total_row].height = 18
 
    c = ws.cell(row=total_row, column=1, value="TOTAL")
    c.font      = total_font
    c.fill      = total_fill
    c.alignment = center
    c.border    = border
 
    for mode in modes:
        sc, ac = col_map[mode]
        total_success = len(logs[mode])
        all_steps     = [e['steps'] for e in logs[mode]]
        overall_avg   = round(np.mean(all_steps), 1) if all_steps else 0.0
 
        for col, val in [(sc, total_success), (ac, overall_avg)]:
            c = ws.cell(row=total_row, column=col, value=val)
            c.font      = total_font
            c.fill      = total_fill
            c.alignment = center
            c.border    = border
 
    ws.freeze_panes = ws.cell(row=data_start_row, column=2)
 
    wb.save(output_path)
    print(f"\nExcel summary saved: {output_path}")
    return output_path
 

if __name__ == "__main__":
    MODES       = ["random", "baseline", "control"]
    MAP_SEED    = 2000
    TOTAL_STEPS = 100000
    CONTROL_SCALE = 2.0
 
    for mode in MODES:
        train_phase2(
            mode=mode,
            control_scale=CONTROL_SCALE,
            map_seed=MAP_SEED,
            total_steps=TOTAL_STEPS,
        )

    export_summary_excel(
        modes=MODES,
        total_steps=TOTAL_STEPS,
        log_freq=2000,
        output_path="phase2_summary.xlsx",
    )